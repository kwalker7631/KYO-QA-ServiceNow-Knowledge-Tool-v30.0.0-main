import pytest
from pathlib import Path

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

if pd is None or load_workbook is None:
    pytest.skip("Required libraries not installed", allow_module_level=True)

from excel_generator import generate_excel
from config import META_COLUMN_NAME, QA_NUMBERS_COLUMN_NAME

DEFAULT_TEMPLATE_HEADERS = [
    "Active",
    "Article type",
    "Author",
    "Category(category)",
    "Configuration item",
    "Confidence",
    "Description",
    "Attachment link",
    "Disable commenting",
    "Disable suggesting",
    "Display attachments",
    "Flagged",
    "Governance",
    "Category(kb_category)",
    "Knowledge Base",
    "Meta",
    "Meta Description",
    "Ownership Group",
    "Published",
    "Scheduled publish date",
    "Short description",
    "Article body",
    "Topic",
    "Problem Code",
    "Product Description",
    "Ticket#",
    "Valid to",
    "View as allowed",
    "Wiki",
    "Sys ID",
    "Process Status",
    "Needs Review",
]

def test_generate_excel_headers(tmp_path):
    sample_df = pd.DataFrame([{"Short description": "Test", "Article body": "body"}])
    output_file = tmp_path / "out.xlsx"
    generate_excel(sample_df.to_dict("records"), output_file)
    assert output_file.exists()

    wb = load_workbook(output_file)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == DEFAULT_TEMPLATE_HEADERS


def test_generate_excel_with_template(tmp_path):
    sample_df = pd.DataFrame([{"Short description": "Test", "Article body": "body"}])
    output_file = tmp_path / "out_template.xlsx"
    template = Path("Sample_Set/kb_knowledge_Template.xlsx")
    generate_excel(sample_df.to_dict("records"), output_file, template)
    assert output_file.exists()

    wb = load_workbook(output_file)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == DEFAULT_TEMPLATE_HEADERS
    assert ws.max_row == 2
    sd_index = DEFAULT_TEMPLATE_HEADERS.index("Short description") + 1
    assert ws.cell(row=2, column=sd_index).value == "Test"

def test_qa_numbers_written_when_column_present(tmp_path):
    sample_df = pd.DataFrame([
        {
            "file_name": "test.pdf",
            "Short description": "test",
            "qa_numbers": "QA-1",
        }
    ])

    template = tmp_path / "template.xlsx"
    wb = load_workbook(Path("Sample_Set/kb_knowledge_Template.xlsx")) if Path("Sample_Set/kb_knowledge_Template.xlsx").exists() else None
    if wb is None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Short description", QA_COLUMN_NAME])
        ws.append(["test", ""])
    wb.save(template)

    output_file = tmp_path / "out.xlsx"
    generate_excel(sample_df.to_dict("records"), output_file, template)

    wb2 = load_workbook(output_file)
    ws2 = wb2.active
    qa_idx = [c.value for c in ws2[1]].index(QA_COLUMN_NAME) + 1
    assert ws2.cell(row=2, column=qa_idx).value == "QA-1"


def test_qa_numbers_not_written_without_column(tmp_path):
    sample_df = pd.DataFrame([
        {"file_name": "test.pdf", "Short description": "test", "qa_numbers": "QA-1"}
    ])

    # Create template without QA column
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Short description"])  # only description
    ws.append(["test"])
    template = tmp_path / "template_noqa.xlsx"
    wb.save(template)

    output_file = tmp_path / "out_noqa.xlsx"
    generate_excel(sample_df.to_dict("records"), output_file, template)

    wb2 = load_workbook(output_file)
    ws2 = wb2.active
    assert QA_COLUMN_NAME not in [c.value for c in ws2[1]]
