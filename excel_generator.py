# excel_generator.py - Definitive version with true cloning, styling, and robust data handling.
# This version correctly preserves all data from the original template.
import logging
logging.getLogger(__name__).setLevel(logging.DEBUG)

try:
    import pandas as pd
except Exception:  # pragma: no cover - library optional in test env
    pd = None
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - library optional in test env
    openpyxl = None
    class _Dummy:
        def __init__(self, *a, **k):
            pass

    Font = PatternFill = Alignment = _Dummy

    def get_column_letter(i):
        return "A"
import re
from pathlib import Path
import shutil

from logging_utils import setup_logger, log_info, log_warning, log_error
from custom_exceptions import ExcelGenerationError
import importlib, sys

try:
    from config import (
        META_COLUMN_NAME,
        AUTHOR_COLUMN_NAME,
        DESCRIPTION_COLUMN_NAME,
        QA_NUMBERS_COLUMN_NAME,
        STATUS_COLUMN_NAME,
    )
except Exception:  # pragma: no cover - allow stubs in tests
    _cfg = importlib.import_module('config')
    META_COLUMN_NAME = getattr(_cfg, 'META_COLUMN_NAME', 'Meta')
    AUTHOR_COLUMN_NAME = getattr(_cfg, 'AUTHOR_COLUMN_NAME', 'Author')
    DESCRIPTION_COLUMN_NAME = getattr(_cfg, 'DESCRIPTION_COLUMN_NAME', 'Short description')
    QA_NUMBERS_COLUMN_NAME = getattr(_cfg, 'QA_NUMBERS_COLUMN_NAME', 'QA Numbers')
    STATUS_COLUMN_NAME = getattr(_cfg, 'STATUS_COLUMN_NAME', 'Processing Status')
    for n, v in (
        ('META_COLUMN_NAME', META_COLUMN_NAME),
        ('AUTHOR_COLUMN_NAME', AUTHOR_COLUMN_NAME),
        ('DESCRIPTION_COLUMN_NAME', DESCRIPTION_COLUMN_NAME),
        ('QA_NUMBERS_COLUMN_NAME', QA_NUMBERS_COLUMN_NAME),
        ('STATUS_COLUMN_NAME', STATUS_COLUMN_NAME),
    ):
        setattr(_cfg, n, v)
    sys.modules['config'] = _cfg

logger = setup_logger("excel_generator")

ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# --- Style Definitions - Restored to match original flair ---
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
CELL_FONT = Font(name="Calibri", size=11)
STATUS_FILLS = {
    "Success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Needs Review": PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    ),
    "Failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Protected": PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    ),
    "Corrupted": PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ),
    "OCR Failed": PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ),
    "No Text Found": PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ),
}

DEFAULT_TEMPLATE_PATH = Path("Sample_Set/kb_knowledge_Template.xlsx")

def _load_default_headers(path=DEFAULT_TEMPLATE_PATH):
    try:
        import zipfile, xml.etree.ElementTree as ET
        with zipfile.ZipFile(path) as z:
            shared = ET.fromstring(z.read("xl/sharedStrings.xml"))
            strings = [t.text for t in shared.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')]
            sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))
            ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            first_row = sheet.find('.//m:sheetData/m:row', ns)
            headers = []
            for c in first_row:
                v = c.find('m:v', ns)
                if v is None:
                    headers.append("")
                elif c.get('t') == 's':
                    headers.append(strings[int(v.text)])
                else:
                    headers.append(v.text)
            return headers
    except Exception:
        return []

DEFAULT_TEMPLATE_HEADERS = _load_default_headers()

# Basic headers used when generating a new template workbook
DEFAULT_TEMPLATE_HEADERS = [
    DESCRIPTION_COLUMN_NAME,
    "Article body",
    META_COLUMN_NAME,
    AUTHOR_COLUMN_NAME,
    QA_NUMBERS_COLUMN_NAME,
    STATUS_COLUMN_NAME,
]

class ExcelWriter:
    """Minimal Excel writer used for unit tests."""

    def __init__(self, file_path: str, headers=None):
        self.file_path = file_path
        self.headers = headers or []
        self.rows = []

    def add_row(self, row: dict):
        self.rows.append(row)

    def save(self):
        with open(self.file_path, "w", encoding="utf-8") as f:
            if self.headers:
                f.write(",".join(self.headers) + "\n")
            for row in self.rows:
                f.write(",".join(str(row.get(h, "")) for h in self.headers) + "\n")


def sanitize_for_excel(value):
    """Sanitizes a value to be safely written to an Excel cell."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def apply_styles(worksheet):
    """Applies all formatting and conditional coloring."""
    log_info(logger, "Applying professional formatting and styles...")

    header_row = worksheet[1]
    for cell in header_row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = CELL_FONT
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )

    for i, column in enumerate(worksheet.columns, 1):
        max_length = 0
        for cell in column:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[get_column_letter(i)].width = min(
            adjusted_width, 70
        )


def generate_excel(all_results, output_path, template_path=Path("Sample_Set/kb_knowledge_Template.xlsx")):
    """
    Generates a formatted Excel file by cloning the template and merging new data.
    """
    try:
        if pd is None or openpyxl is None:
            raise ExcelGenerationError("Required libraries not installed")

        if not all_results:
            raise ExcelGenerationError(
                "No data was processed to generate an Excel file."
            )

        new_data_df = pd.DataFrame(all_results).applymap(sanitize_for_excel)
        cols = getattr(new_data_df, 'columns', [])
        if 'qa_numbers' in cols and QA_NUMBERS_COLUMN_NAME not in cols:
            new_data_df.rename(columns={'qa_numbers': QA_NUMBERS_COLUMN_NAME}, inplace=True)
        new_data_df['merge_key'] = new_data_df['file_name'].apply(lambda x: Path(x).stem)
        new_data_df.set_index('merge_key', inplace=True)


        if not template_path.exists():
            raise ExcelGenerationError(f"Template file not found at: {template_path}")

        shutil.copy(template_path, output_path)

        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active
        header = [cell.value for cell in worksheet[1]]
        desc_col_idx = header.index(DESCRIPTION_COLUMN_NAME) if DESCRIPTION_COLUMN_NAME in header else -1
        meta_col_idx = header.index(META_COLUMN_NAME) if META_COLUMN_NAME in header else -1
        author_col_idx = header.index(AUTHOR_COLUMN_NAME) if AUTHOR_COLUMN_NAME in header else -1
        qa_col_idx = header.index(QA_NUMBERS_COLUMN_NAME) if QA_NUMBERS_COLUMN_NAME in header else -1

        if desc_col_idx == -1:
            raise ExcelGenerationError(
                f"Template missing required column: '{DESCRIPTION_COLUMN_NAME}'"
            )

        for row_num, row_cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
            desc_val = row_cells[desc_col_idx].value
            if not desc_val:
                continue

            # Allow template short descriptions prefixed with "Processed:" (case-insensitive) to
            # merge correctly with new data indexed by file name.
            row_text = re.sub(r"^Processed:\s*", "", str(desc_val), flags=re.IGNORECASE)
            row_key = Path(row_text).stem
            if row_key in new_data_df.index:
                new_row_data = new_data_df.loc[row_key]

                if meta_col_idx != -1 and META_COLUMN_NAME in new_row_data:
                    worksheet.cell(row=row_num, column=meta_col_idx + 1).value = new_row_data[META_COLUMN_NAME]

                if qa_col_idx != -1 and "qa_numbers" in new_row_data and new_row_data["qa_numbers"]:
                    worksheet.cell(row=row_num, column=qa_col_idx + 1).value = new_row_data["qa_numbers"]
                elif meta_col_idx != -1 and "qa_numbers" in new_row_data and new_row_data["qa_numbers"]:
                    existing = worksheet.cell(row=row_num, column=meta_col_idx + 1).value or ""
                    sep = " | " if existing else ""
                    worksheet.cell(row=row_num, column=meta_col_idx + 1).value = f"{existing}{sep}{new_row_data['qa_numbers']}"

                if author_col_idx != -1 and AUTHOR_COLUMN_NAME in new_row_data:
                    worksheet.cell(row=row_num, column=author_col_idx + 1).value = (
                        new_row_data[AUTHOR_COLUMN_NAME]
                    )

                if fill := STATUS_FILLS.get(new_row_data["processing_status"]):
                    for cell in row_cells:
                        cell.fill = fill
            else:
                log_warning(
                    logger,
                    f"Skipped template row {row_num} ({desc_val}) - no processed data found",
                )

        apply_styles(worksheet)
        workbook.save(output_path)

        log_info(
            logger, f"Successfully created cloned and updated Excel file: {output_path}"
        )
        return str(output_path)

    except Exception as e:
        log_error(logger, f"Excel generation failed: {e}")
        raise ExcelGenerationError(f"Failed to generate Excel file: {e}")
