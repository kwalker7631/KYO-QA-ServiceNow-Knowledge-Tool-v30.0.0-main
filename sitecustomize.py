# Ensure project root is in sys.path for imports
import sys
from pathlib import Path

ROOT = Path(__file__).parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Helper function to create and register mock modules
def create_mock_module(name, attributes):
    import types, sys
    module = types.ModuleType(name)
    for attr_name, attr_value in attributes.items():
        setattr(module, attr_name, attr_value)
    sys.modules.setdefault(name, module)
    return module

# Ensure openpyxl stub is available for unit tests
try:  # pragma: no cover - optional dependency
    import openpyxl  # noqa: F401
except Exception:
    openpyxl_styles = {
        "Alignment": type("Alignment", (), {}),
        "Font": type("Font", (), {}),
        "PatternFill": type("PatternFill", (), {"__init__": lambda self, *a, **k: None}),
    }
    openpyxl_utils = {
        "get_column_letter": lambda x: "A",
        "dataframe": create_mock_module("dataframe", {"dataframe_to_rows": lambda df: []}),
    }
    openpyxl_worksheet = {
        "copier": create_mock_module("copier", {"WorksheetCopy": object}),
    }
    openpyxl = create_mock_module("openpyxl", {
        "styles": create_mock_module("styles", openpyxl_styles),
        "utils": create_mock_module("utils", openpyxl_utils),
        "worksheet": create_mock_module("worksheet", openpyxl_worksheet),
    })

# Provide a lightweight pandas stub for the test environment
try:  # pragma: no cover - only for tests
    import pandas  # type: ignore  # noqa: F401
except Exception:  # pragma: no cover
    class _DF(list):
        def to_dict(self, *_, **__):
            return list(self)

    pandas = create_mock_module("pandas", {
        "DataFrame": lambda data: _DF(data),
    })
