# processing_engine.py - Enhanced processing engine with comprehensive error handling
import shutil, time, json, openpyxl, re
from queue import Queue
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import *
from custom_exceptions import FileLockError, PDFProtectionError, PDFCorruptionError
from data_harvesters import harvest_all_data
from file_utils import is_file_locked
from ocr_utils import process_single_document, _is_ocr_needed

def clear_review_folder():
    """Clear any existing review files."""
    if PDF_TXT_DIR.exists():
        for f in PDF_TXT_DIR.glob("*.txt"):
            try:
                f.unlink()
            except OSError as e:
                print(f"Error deleting review file {f}: {e}")

def get_cache_path(pdf_path):
    """Generate cache file path for a PDF."""
    try:
        return CACHE_DIR / f"{pdf_path.stem}_{pdf_path.stat().st_size}.json"
    except FileNotFoundError:
        return CACHE_DIR / f"{pdf_path.stem}_unknown.json"

def process_single_pdf(pdf_path, progress_queue, ignore_cache=False):
    """
    Process a single PDF file with enhanced error handling and detailed reporting.
    
    Args:
        pdf_path: Path to the PDF file
        progress_queue: Queue for progress updates
        ignore_cache: Whether to skip cache lookup
        
    Returns:
        dict: Comprehensive processing result with status, failure reason, and extracted data
    """
    pdf_path = Path(pdf_path)
    filename = pdf_path.name
    cache_path = get_cache_path(pdf_path)

    # Announce processing start
    progress_queue.put({
        "type": "log", 
        "tag": "info", 
        "msg": f"Processing: {filename}"
    })
    
    # Check cache first (unless ignoring)
    if not ignore_cache and cache_path.exists():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            
            # Validate cache data structure
            if all(key in cached_data for key in ["status", "filename", "models"]):
                progress_queue.put({
                    "type": "log", 
                    "tag": "info", 
                    "msg": f"Loaded from cache: {filename}"
                })
                
                # Handle cached review items
                if cached_data.get("status") == "Needs Review" and cached_data.get("review_info"):
                    progress_queue.put({
                        "type": "review_item", 
                        "data": cached_data.get("review_info")
                    })
                
                # Report completion
                progress_queue.put({
                    "type": "file_complete", 
                    "status": cached_data.get("status")
                })
                
                if cached_data.get("ocr_used"):
                    progress_queue.put({
                        "type": "increment_counter", 
                        "counter": "ocr"
                    })
                
                return cached_data
                
        except (json.JSONDecodeError, KeyError) as e:
            progress_queue.put({
                "type": "log", 
                "tag": "warning", 
                "msg": f"Corrupt cache for {filename}. Reprocessing..."
            })

    # Update status to show processing
    progress_queue.put({
        "type": "status", 
        "msg": f"Processing {filename}", 
        "led": "Processing"
    })
    
    # Initialize result structure
    result = {
        "filename": filename,
        "models": "",
        "author": "",
        "status": "Fail",
        "failure_reason": "",
        "ocr_used": False,
        "review_info": None,
        "processing_time": 0
    }
    
    start_time = time.time()
    
    try:
        # Check if OCR will be needed (for progress reporting)
        try:
            ocr_required = _is_ocr_needed(str(pdf_path.resolve()))
            result["ocr_used"] = ocr_required
            
            if ocr_required:
                progress_queue.put({
                    "type": "status", 
                    "msg": f"OCR required for {filename}", 
                    "led": "OCR"
                })
                progress_queue.put({
                    "type": "increment_counter", 
                    "counter": "ocr"
                })
        except Exception as e:
            # If we can't determine OCR needs, assume it might be needed
            result["ocr_used"] = True
            progress_queue.put({
                "type": "log", 
                "tag": "warning", 
                "msg": f"Could not determine OCR needs for {filename}: {e}"
            })
        
        # Process the document using the new comprehensive function
        status, failure_reason, extracted_text = process_single_document(pdf_path)
        
        # Map processing status to our internal statuses
        status_mapping = {
            "success": "Pass",
            "protected": "Protected", 
            "corrupted": "Corrupted",
            "ocr_failed": "OCR Failed",
            "no_text": "No Text",
            "error": "Fail"
        }
        
        mapped_status = status_mapping.get(status, "Fail")
        result["status"] = mapped_status
        result["failure_reason"] = failure_reason or ""
        
        if status == "success" and extracted_text:
            # Text extraction successful, now extract metadata
            progress_queue.put({
                "type": "status", 
                "msg": f"Extracting data from {filename}", 
                "led": "AI"
            })
            
            try:
                data = harvest_all_data(extracted_text, filename)
                result["models"] = data["models"]
                result["author"] = data["author"]
                
                if data["models"] == "Not Found":
                    # Models not found - needs review
                    result["status"] = "Needs Review"
                    result["failure_reason"] = "No model patterns matched in document text"
                    
                    # Create review file
                    review_txt_path = PDF_TXT_DIR / f"{pdf_path.stem}.txt"
                    with open(review_txt_path, 'w', encoding='utf-8') as f:
                        f.write(f"--- Filename: {filename} ---\n")
                        f.write(f"--- Status: {result['status']} ---\n")
                        f.write(f"--- Reason: {result['failure_reason']} ---\n")
                        f.write(f"--- OCR Used: {result['ocr_used']} ---\n\n")
                        f.write("=== EXTRACTED TEXT ===\n\n")
                        f.write(extracted_text)
                    
                    review_info = {
                        "filename": filename,
                        "reason": result["failure_reason"],
                        "txt_path": str(review_txt_path),
                        "pdf_path": str(pdf_path),
                        "ocr_used": result["ocr_used"],
                        "status": result["status"]
                    }
                    result["review_info"] = review_info
                    
                    progress_queue.put({
                        "type": "review_item", 
                        "data": review_info
                    })
                    
                    progress_queue.put({
                        "type": "log", 
                        "tag": "warning", 
                        "msg": f"Needs review: {filename} - No model patterns found"
                    })
                else:
                    # Success!
                    result["status"] = "Pass"
                    result["failure_reason"] = ""
                    progress_queue.put({
                        "type": "log", 
                        "tag": "info", 
                        "msg": f"Success: {filename} - Found models: {data['models']}"
                    })
                    
            except Exception as e:
                # Data harvesting failed
                result["failure_reason"] = f"Pattern matching failed: {str(e)}"
                result["models"] = "Error: Pattern Matching Failed"
                result["status"] = "Fail"
                progress_queue.put({
                    "type": "log", 
                    "tag": "error", 
                    "msg": f"Pattern matching failed for {filename}: {e}"
                })
        
        else:
            # Text extraction failed
            if status == "protected":
                result["models"] = "Error: Protected PDF"
                progress_queue.put({
                    "type": "log", 
                    "tag": "warning", 
                    "msg": f"Protected PDF: {filename} - {failure_reason}"
                })
            elif status == "corrupted":
                result["models"] = "Error: Corrupted PDF"
                progress_queue.put({
                    "type": "log", 
                    "tag": "error", 
                    "msg": f"Corrupted PDF: {filename} - {failure_reason}"
                })
            elif status in ["ocr_failed", "no_text"]:
                result["models"] = "Error: Text Extraction Failed"
                progress_queue.put({
                    "type": "log", 
                    "tag": "warning", 
                    "msg": f"Text extraction failed: {filename} - {failure_reason}"
                })
            else:
                result["models"] = "Error: Processing Failed"
                progress_queue.put({
                    "type": "log", 
                    "tag": "error", 
                    "msg": f"Processing failed: {filename} - {failure_reason}"
                })

    except Exception as e:
        # Catch-all for any unexpected errors
        result["failure_reason"] = f"Unexpected error: {str(e)}"
        result["status"] = "Fail"
        result["models"] = "Error: Critical Failure"
        progress_queue.put({
            "type": "log", 
            "tag": "error", 
            "msg": f"Critical error processing {filename}: {e}"
        })

    # Record processing time
    result["processing_time"] = time.time() - start_time
    
    # Cache the result
    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2)
    except Exception as e:
        progress_queue.put({
            "type": "log", 
            "tag": "warning", 
            "msg": f"Failed to cache result for {filename}: {e}"
        })
    
    # Report completion
    progress_queue.put({
        "type": "file_complete", 
        "status": result["status"]
    })
    
    return result

def run_processing_job(job_info, progress_queue, cancel_event, pause_event):
    """
    Enhanced processing job with comprehensive error handling and progress reporting.
    
    Args:
        job_info: Dictionary containing job configuration
        progress_queue: Queue for UI updates
        cancel_event: Event to signal job cancellation
        pause_event: Event to signal job pause
    """
    try:
        is_rerun = job_info.get("is_rerun", False)
        excel_path = Path(job_info["excel_path"])
        input_path = job_info["input_path"]
        
        progress_queue.put({
            "type": "log",
            "tag": "info",
            "msg": f"Processing job started. Rerun: {is_rerun}"
        })
        total_steps = 5  # Define the total number of steps in the workflow
        current_step = 1  # Current step: Initial progress
        progress_queue.put({"type": "import_progress", "value": int((current_step / total_steps) * 100)})

        # Handle Excel file cloning
        if is_rerun:
            clear_review_folder()
            cloned_path = excel_path
            progress_queue.put({
                "type": "log", 
                "tag": "info", 
                "msg": "Rerun mode: Using existing Excel file"
            })
        else:
            ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            cloned_path = OUTPUT_DIR / f"cloned_{excel_path.stem}_{ts}{excel_path.suffix}"
            
            # Check if Excel file is locked
            if is_file_locked(excel_path):
                raise FileLockError(f"Input Excel file is locked: {excel_path.name}")
            
            # Clone the Excel file
            shutil.copy(excel_path, cloned_path)
            progress_queue.put({
                "type": "log",
                "tag": "info",
                "msg": f"Excel file cloned to: {cloned_path.name}"
            })
            progress_queue.put({"type": "import_progress", "value": 50})
        
        # Get list of files to process
        if isinstance(input_path, list):
            files = [Path(f) for f in input_path]
        else:
            files = list(Path(input_path).glob('*.pdf'))
        
        total_files = len(files)
        
        if total_files == 0:
            progress_queue.put({
                "type": "log", 
                "tag": "warning", 
                "msg": "No PDF files found to process"
            })
            progress_queue.put({
                "type": "finish", 
                "status": "No Files"
            })
            return
        
        progress_queue.put({
            "type": "log", 
            "tag": "info", 
            "msg": f"Found {total_files} PDF files to process"
        })
        
        # Process files
        results = {}
        processed_count = 0
        start_time = time.time()
        
        for i, path in enumerate(files):
            # Check for cancellation
            if cancel_event.is_set():
                progress_queue.put({
                    "type": "log", 
                    "tag": "warning", 
                    "msg": "Processing cancelled by user"
                })
                break
                
            # Handle pause
            if pause_event and pause_event.is_set():
                progress_queue.put({
                    "type": "status", 
                    "msg": "Processing paused", 
                    "led": "Paused"
                })
                while pause_event.is_set():
                    if cancel_event.is_set():
                        break
                    time.sleep(0.5)
                if cancel_event.is_set():
                    break
                progress_queue.put({
                    "type": "status", 
                    "msg": "Processing resumed", 
                    "led": "Processing"
                })
            
            # Update progress
            processed_count = i + 1
            progress_queue.put({
                "type": "progress", 
                "current": processed_count, 
                "total": total_files
            })
            
            # Process the file
            try:
                result = process_single_pdf(path, progress_queue, ignore_cache=is_rerun)
                if result:
                    results[result["filename"]] = result
                    
            except Exception as e:
                progress_queue.put({
                    "type": "log", 
                    "tag": "error", 
                    "msg": f"Critical error processing {path.name}: {e}"
                })
                
                # Create a failure result for this file
                results[path.name] = {
                    "filename": path.name,
                    "models": "Error: Critical Failure",
                    "author": "",
                    "status": "Fail",
                    "failure_reason": f"Critical processing error: {str(e)}",
                    "ocr_used": False,
                    "review_info": None,
                    "processing_time": 0
                }

        # Check for cancellation before Excel update
        if cancel_event.is_set():
            progress_queue.put({
                "type": "finish", 
                "status": "Cancelled"
            })
            return

        # Update Excel file with results
        progress_queue.put({
            "type": "status", 
            "msg": "Updating Excel file...", 
            "led": "Saving"
        })
        progress_queue.put({
            "type": "log", 
            "tag": "info", 
            "msg": f"Updating Excel file with {len(results)} results"
        })
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(cloned_path)
            progress_queue.put({"type": "import_progress", "value": 75})
            sheet = workbook.active
            headers = [c.value for c in sheet[1]]
            
            # Add status column if it doesn't exist
            if STATUS_COLUMN_NAME not in headers:
                sheet.cell(row=1, column=len(headers) + 1).value = STATUS_COLUMN_NAME
                headers.append(STATUS_COLUMN_NAME)
            
            # Add failure reason column if it doesn't exist
            failure_reason_col = "Failure Reason"
            if failure_reason_col not in headers:
                sheet.cell(row=1, column=len(headers) + 1).value = failure_reason_col
                headers.append(failure_reason_col)
            
            # Create column mapping
            required_columns = [
                DESCRIPTION_COLUMN_NAME, 
                META_COLUMN_NAME, 
                AUTHOR_COLUMN_NAME, 
                STATUS_COLUMN_NAME, 
                failure_reason_col
            ]
            
            cols = {}
            for col_name in required_columns:
                if col_name in headers:
                    cols[col_name] = headers.index(col_name) + 1
                else:
                    progress_queue.put({
                        "type": "log", 
                        "tag": "warning", 
                        "msg": f"Column '{col_name}' not found in Excel file"
                    })
           
            # Update rows with results
            rows_updated = 0
            for row in sheet.iter_rows(min_row=2):
                if DESCRIPTION_COLUMN_NAME not in cols:
                    continue
                    
                desc = str(row[cols[DESCRIPTION_COLUMN_NAME]-1].value or "")
                
                for filename, data in results.items():
                    if Path(filename).stem in desc:
                        # Update cells
                        if META_COLUMN_NAME in cols:
                            row[cols[META_COLUMN_NAME]-1].value = data["models"]
                        if AUTHOR_COLUMN_NAME in cols:
                            row[cols[AUTHOR_COLUMN_NAME]-1].value = data["author"]
                        if STATUS_COLUMN_NAME in cols:
                            status_text = f"{data['status']}{' (OCR)' if data['ocr_used'] else ''}"
                            row[cols[STATUS_COLUMN_NAME]-1].value = status_text
                        if failure_reason_col in cols:
                            row[cols[failure_reason_col]-1].value = data.get("failure_reason", "")
                        
                        rows_updated += 1
                        break
            
            progress_queue.put({
                "type": "log", 
                "tag": "info", 
                "msg": f"Updated {rows_updated} rows in Excel file"
            })
            
            # Apply formatting
            progress_queue.put({
                "type": "status", 
                "msg": "Applying formatting...", 
                "led": "Saving"
            })
            
            fills = {
                "Pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                "Fail": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                "Needs Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                "Protected": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                "Corrupted": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
                "OCR Failed": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
                "No Text": PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),
                "OCR": PatternFill(start_color="0A9BCD", end_color="0A9BCD", fill_type="solid")
            }
            
            # Apply conditional formatting
            if STATUS_COLUMN_NAME in cols:
                for row in sheet.iter_rows(min_row=2):
                    status_val = str(row[cols[STATUS_COLUMN_NAME]-1].value or "")
                    fill_key = status_val.replace(" (OCR)", "").strip()
                    fill = fills.get(fill_key)
                    
                    if fill:
                        for cell in row:
                            cell.fill = fill
                    
                    # Special highlighting for OCR
                    if "(OCR)" in status_val and "OCR" in fills:
                        row[cols[STATUS_COLUMN_NAME]-1].fill = fills["OCR"]
            
            # Auto-adjust column widths
            for i, col in enumerate(sheet.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                sheet.column_dimensions[get_column_letter(i)].width = min((max_len + 2), 60)

            # Save the workbook
            workbook.save(cloned_path)
            progress_queue.put({"type": "import_progress", "value": 100})
            progress_queue.put({
                "type": "result_path",
                "path": str(cloned_path)
            })
            # Notify UI that the result can now be opened
            progress_queue.put({"type": "enable_open_result"})
            
            # Generate processing summary
            total_time = time.time() - start_time
            status_counts = {}
            
            for result in results.values():
                status = result["status"]
                status_counts[status] = status_counts.get(status, 0) + 1
            
            summary_parts = [f"{status}: {count}" for status, count in status_counts.items()]
            summary_msg = f"Processing complete in {total_time:.1f}s: {processed_count}/{total_files} files processed. {', '.join(summary_parts)}"
            
            progress_queue.put({
                "type": "log", 
                "tag": "success", 
                "msg": summary_msg
            })
            
            progress_queue.put({
                "type": "finish", 
                "status": "Complete"
            })

        except Exception as e:
            progress_queue.put({
                "type": "log", 
                "tag": "error", 
                "msg": f"Failed to update Excel file: {e}"
            })
            progress_queue.put({
                "type": "finish", 
                "status": f"Excel Error: {e}"
            })

    except Exception as e:
        progress_queue.put({
            "type": "log", 
            "tag": "error", 
            "msg": f"Critical job error: {e}"
        })
        progress_queue.put({
            "type": "finish", 
            "status": f"Error: {e}"
        })